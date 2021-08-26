import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Alignment

import datetime
class savingMeasurements(object):
    def __init__(self, UserNumber,TestNumber,NumberOfCars):
        self.NumberOfCars=NumberOfCars
        self.x=3
        self.FilePathName='./ExportData/Test '+str(TestNumber)+' User '+str(UserNumber)+'.xlsx'

        wb = openpyxl.Workbook()

        PowerWorksheet = wb.create_sheet("PowerMeausurments")
        EnergyWorksheet= wb.create_sheet("EnergyMeausurments")

        self.fontStyleWord = Font(name="Calibri",size = "8")
        self.fontStyleNumber=Font(name="Calibri",size = "10")

        self.alignmentStyle=Alignment(horizontal='center',vertical='center')


        EnergyWorksheet.cell(row = self.x, column = 2, value = 'Time').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 3, value = 'Wallet[€]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 4, value = 'Price[€]').font = self.fontStyleWord

        EnergyWorksheet.cell(row = self.x, column = 5, value = 'Ein[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 6, value = 'Eout[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 7, value = 'EdSr[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 8, value = 'EdLd[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 9, value = 'EbAvSr[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 10, value = 'EbAvLd[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 11, value = 'EbRqLd[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 12, value = 'Ehsb[kW/h]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 13, value = 'SOChsb[%]').font = self.fontStyleWord


        PowerWorksheet.cell(row = self.x, column = 2, value = 'Time').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 3, value = 'Wallet[€]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 4, value = 'Price[€]').font = self.fontStyleWord

        PowerWorksheet.cell(row = self.x, column = 5, value = 'Pout[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 6, value = 'Pin[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 7, value = 'PdSr[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 8, value = 'PdLd[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 9, value = 'PdAvSr[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 10, value = 'PdAvLd[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 11, value = 'PdRqLd[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 12, value = 'Phsb[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 13, value = 'SOChsb[%]').font = self.fontStyleWord

        numberOfCell=13

        if NumberOfCars<0:

            for q in range (self.NumberOfCars):
                EnergyWorksheet.cell(row = self.x, column = 14+3*q, value = 'Ecar'+str(self.NumberOfCars)+'[kWh]').font = self.fontStyleWord
                EnergyWorksheet.cell(row = self.x, column = 15+3*q, value = 'SOCcar'+str(self.NumberOfCars)+'[%]').font = self.fontStyleWord

                PowerWorksheet.cell(row = self.x, column = 14+3*q, value = 'Pcar'+str(self.NumberOfCars)+'[kW]').font = self.fontStyleWord
                PowerWorksheet.cell(row = self.x, column = 15+3*q, value = 'SOCcar'+str(self.NumberOfCars)+'[%]').font = self.fontStyleWord
                numberOfCell+=3


        for q in range (2,numberOfCell+1):
            EnergyWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            PowerWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            colume= get_column_letter(q)
            EnergyWorksheet.column_dimensions[colume].width =9.4
            PowerWorksheet.column_dimensions[colume].width =9.4


        EnergyWorksheet.column_dimensions["B"].width =15
        PowerWorksheet.column_dimensions["B"].width =15


        wb.save(filename = self.FilePathName)
        wb.close()

        wb = openpyxl.load_workbook(filename =self.FilePathName)
        Sheetworksheet= wb["Sheet"]
        wb.remove(Sheetworksheet)
        wb.save(filename = self.FilePathName)
        wb.close()

    def safeBasicMeasurements(self,DataTime,AvgPout,AvgPin,AvgArrTotPower,SumEout,SumEin,SumArrTotEnergy):

        self.x+=1
        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        k=1000 #convert W to kW
        h=3600 #convert Ws to Wh

        date = datetime.datetime.strptime(DataTime, "%d/%m/%Y %H:%M")

        EnergyWorksheet.cell(row = self.x, column = 2, value = date).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 2, value = date).font= self.fontStyleNumber

        EnergyWorksheet.cell(row = self.x, column = 2, value = date).number_format = 'DD/MM/YYYY HH:MM'
        PowerWorksheet.cell(row = self.x, column = 2, value = date).number_format = 'DD/MM/YYYY HH:MM'


        EnergyWorksheet.cell(row = self.x, column = 5, value = round(SumEout/(k*h),3))
        EnergyWorksheet.cell(row = self.x, column = 6, value = round(SumEin/(k*h),3)) 
        EnergyWorksheet.cell(row = self.x, column = 7, value = round(SumArrTotEnergy[0]/(k*h),3))
        EnergyWorksheet.cell(row = self.x, column = 8, value = round(SumArrTotEnergy[1]/(k*h),3))
        EnergyWorksheet.cell(row = self.x, column = 9, value = round(SumArrTotEnergy[2]/(k*h),3))
        EnergyWorksheet.cell(row = self.x, column = 10, value = round(SumArrTotEnergy[3]/(k*h),3))
        EnergyWorksheet.cell(row = self.x, column = 11, value = round(SumArrTotEnergy[4]/(k*h),3))

        PowerWorksheet.cell(row = self.x, column = 5, value = round(AvgPout/k,3))
        PowerWorksheet.cell(row = self.x, column = 6, value = round(AvgPin/k,3))
        PowerWorksheet.cell(row = self.x, column = 7, value = round(AvgArrTotPower[0]/k,3))
        PowerWorksheet.cell(row = self.x, column = 8, value = round(AvgArrTotPower[1]/k,3))
        PowerWorksheet.cell(row = self.x, column = 9, value = round(AvgArrTotPower[2]/k,3))
        PowerWorksheet.cell(row = self.x, column = 10, value = round(AvgArrTotPower[3]/k,3))
        PowerWorksheet.cell(row = self.x, column = 11, value = round(AvgArrTotPower[4]/k,3))

        for q in range (2,12):

            PowerWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            PowerWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            if q>2:
                EnergyWorksheet.cell(row = self.x, column = q).number_format = '#,##0.000'
                PowerWorksheet.cell(row = self.x, column = q).number_format = '#,##0.000'

        wb.save(filename = self.FilePathName)
        wb.close()



    def safeHomeBatteryMeasurements(self,InfoBat):

        k=1000 #convert W to kW
        h=3600 #convert Ws to Wh
        p=100
        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        PowerWorksheet.cell(row = self.x, column = 12, value = InfoBat[0]/k)
        PowerWorksheet.cell(row = self.x, column = 13, value = InfoBat[2]*p)

        EnergyWorksheet.cell(row = self.x, column = 12, value = InfoBat[1]/(k*h))
        EnergyWorksheet.cell(row = self.x, column = 13, value = InfoBat[2]*p)


        for q in range (12,14):
            
            PowerWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            PowerWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            if q==14:
                EnergyWorksheet.cell(row = self.x, column = q).number_format = '#,##0.0'
                PowerWorksheet.cell(row = self.x, column = q).number_format = '#,##0.0'
            else:
                EnergyWorksheet.cell(row = self.x, column = q).number_format = '#,##0.000'
                PowerWorksheet.cell(row = self.x, column = q).number_format = '#,##0.000'

        wb.save(filename = self.FilePathName)
        wb.close()

    def safeCarBatteryMeasurements(self,CarNumber,InfoBat):

        k=1000 #convert W to kW
        h=3600 #convert Ws to Wh
        p=100
        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        PowerWorksheet.cell(row = self.x, column = 14+(3*CarNumber), value=InfoBat[0]/k)
        PowerWorksheet.cell(row = self.x, column = 15+(3*CarNumber), value = InfoBat[2]*p)

        EnergyWorksheet.cell(row = self.x, column = 14+(3*CarNumber), value = InfoBat[1]/(k*h))
        EnergyWorksheet.cell(row = self.x, column = 15+(3*CarNumber), value = InfoBat[2]*p)

        for q in range (14+(3*CarNumber),16+(3*CarNumber)):
            
            PowerWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            PowerWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle

            if q==1:
                EnergyWorksheet.cell(row = self.x, column = q).number_format = '#,##0.0'
                PowerWorksheet.cell(row = self.x, column = q).number_format = '#,##0.0'
            else:
                EnergyWorksheet.cell(row = self.x, column = q).number_format = '#,##0.000'
                PowerWorksheet.cell(row = self.x, column = q).number_format = '#,##0.000'

        wb.save(filename = self.FilePathName)
        wb.close()


    def safeCashBalance(self,MonayWalletCent, PriceForEnergyCent):

        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        EnergyWorksheet.cell(row = self.x, column = 3, value = MonayWalletCent/100)
        EnergyWorksheet.cell(row = self.x, column = 4, value =PriceForEnergyCent/100)

        PowerWorksheet.cell(row = self.x, column = 3, value = MonayWalletCent/100)
        PowerWorksheet.cell(row = self.x, column = 4, value =PriceForEnergyCent/100)

        for q in range (3,5):

            PowerWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).font = self.fontStyleNumber
            EnergyWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            PowerWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            EnergyWorksheet.cell(row = self.x, column = q).number_format = '#,##0.00'
            PowerWorksheet.cell(row = self.x, column = q).number_format = '#,##0.00'

        wb.save(filename = self.FilePathName)
        wb.close()


    def deletePreviousValues(self):

        wb = openpyxl.load_workbook(filename =self.FilePathName)
        worksheet= wb["SystemMeausurments"]

        for q in (3, self.x + 1):
            for i in range(2, 16+3*self.CarNumber):
                worksheet.delete_rows(q, i)

        wb.save(filename = self.FilePathName)
        wb.close()     