from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import matplotlib.lines as mlines
from matplotlib.ticker import MultipleLocator

import numpy as np

def readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber):

    FolderPath=str(FileDirecotoryUserData)+'Test '+str(TestNumber)+' User '+str(UserNumber)+'.xlsx'

    wb = load_workbook(filename = FolderPath)
    xlsxPowerMeasurments = wb["PowerMeausurments"]

    row=100
    Pin=[0]*97
    Pout=[0]*97
    PdSr=[0]*97
    PdLd=[0]*97
    PbAvSr=[0]*97
    PbAvLd=[0]*97
    PbRqLd=[0]*97
    IntTime=[0]*97

    for q in range(97):
        Pout[q] -= xlsxPowerMeasurments["E"+str(row+q)].value
        Pin[q] = xlsxPowerMeasurments["F"+str(row+q)].value
        PdSr[q] -= xlsxPowerMeasurments["G"+str(row+q)].value
        PdLd[q] = xlsxPowerMeasurments["H"+str(row+q)].value
        PbAvSr[q] -= xlsxPowerMeasurments["I"+str(row+q)].value
        PbAvLd[q] = xlsxPowerMeasurments["J"+str(row+q)].value
        PbRqLd[q] = xlsxPowerMeasurments["K"+str(row+q)].value
        IntTime[q]= xlsxPowerMeasurments["B"+str(row+q)].value

        matrixUserData=[IntTime,Pout,Pin,PdSr,PdLd,PbAvSr,PbAvLd,PbRqLd]

    return(matrixUserData)


def readUserEnergyData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber):

    FolderPath=str(FileDirecotoryUserData)+'Test '+str(TestNumber)+' User '+str(UserNumber)+'.xlsx'

    wb = load_workbook(filename = FolderPath)
    xlsxPowerMeasurments = wb["EnergyMeausurments"]

    row=101
    Price=[0]*97
    Wout=[0]*97
    Win=[0]*97
    IntTime=[0]*97
    PriceKWH=[0]*97

    for q in range(97):

        Price[q]= xlsxPowerMeasurments["D"+str(row+q)].value
        Wout[q] = xlsxPowerMeasurments["E"+str(row+q)].value
        Win[q] = xlsxPowerMeasurments["F"+str(row+q)].value
        IntTime[q]= xlsxPowerMeasurments["B"+str(row+q)].value

        if Wout[q]!=0 or Win[q]!=0:
            PriceKWH[q]=(100*Price[q])/(Wout[q]+Win[q])
        else:
            PriceKWH[q]=0

    print(PriceKWH)
    return(IntTime,PriceKWH)


def drawingPriceGraph4Users(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    fig, axs = plt.subplots(2, 2,figsize=(13,11),constrained_layout=True)

    UserData=[0]*4
    PriceLow=0
    PriceHigh=0

    for q in range (4):
        UserNumber=ArrUserNumber[q]
        UserData[q]=readUserEnergyData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)
        r=UserData[q]

        for y in range (1,2):
            for x in range (len(UserData[q][y])):
                PriceKWH=UserData[q][y][x]
                if PriceLow>PriceKWH:
                    PriceLow=PriceKWH

                if PriceKWH>PriceHigh:
                    PriceHigh=PriceKWH

    for q in range (4):
        if q<2:
            drawingPriceGraph(axs[0][q],UserData[q],PriceLow,PriceHigh)
            axs[0][q].set_title('UPORABNIK '+str(ArrUserNumber[q]),fontsize=12)
        
        else:
            drawingPriceGraph(axs[1][q-2],UserData[q],PriceLow,PriceHigh)
            axs[1][q-2].set_title('UPORABNIK '+str(ArrUserNumber[q]),fontsize=12)


    plt.show()


def drawingPowerGraph4Users(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    fig, axs = plt.subplots(2, 2,figsize=(13,11),constrained_layout=True)

    UserData=[0]*4
    PowerLow=0
    PowerHigh=0

    for q in range (4):
        UserNumber=ArrUserNumber[q]
        UserData[q]=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)
        r=UserData[q]

        for y in range (1,8):
            for x in range (len(UserData[q][y])):
                Power=UserData[q][y][x]
                if PowerLow>Power:
                    PowerLow=Power

                if Power>PowerHigh:
                    PowerHigh=Power

    for q in range (4):
        if q<2:
            drawingPowerGraph(axs[0][q],UserData[q],PowerLow,PowerHigh)
            axs[0][q].set_title('UPORABNIK '+str(ArrUserNumber[q]),fontsize=12)
        
        else:
            drawingPowerGraph(axs[1][q-2],UserData[q],PowerLow,PowerHigh)
            axs[1][q-2].set_title('UPORABNIK '+str(ArrUserNumber[q]),fontsize=12)


        if q==3:
            axs[1][q-2].legend(loc="lower right", borderaxespad=0.5,fontsize=11)

    fig.set_constrained_layout_pads(hspace=0.05)

    plt.show()


def drawingPowerSystemGraph(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    NumberOfUser=len(ArrUserNumber)
    UserData=[0]*NumberOfUser
    PowerLow=0
    PowerHigh=0
    
    for q in range (NumberOfUser):

        UserNumber=ArrUserNumber[q]
        UserData=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)

        if q==0:
            SystemData=UserData

        else:
            for y in range (1,8):
                for x in range (len(UserData[0])):
                    SystemData[y][x]+=UserData[y][x]


    for y in range (1,8):
        for x in range (len(SystemData[0])):
            Power=SystemData[y][x]
            if PowerLow>Power:
                PowerLow=Power
            if Power>PowerHigh:
                PowerHigh=Power
 
    print(PowerLow)

    fig, ax = plt.subplots(figsize=(11,8),constrained_layout=True)
    drawingPowerGraph(ax,SystemData,PowerLow,PowerHigh)
    ax.legend(loc="lower right", borderaxespad=0.5,fontsize=11)
    plt.show()


def drawingPowerGraph(axs,UserData,PowerLow,PowerHigh):


    axs.fill_between(UserData[0],UserData[3],  step="pre", alpha=0.3,facecolor='#1515ff', label="PdSr")
    axs.plot(UserData[0],UserData[3],drawstyle="steps",color='#1515ff',linewidth=0.8, alpha=0.7)

    axs.fill_between(UserData[0],UserData[4],  step="pre", alpha=0.3,facecolor='red', label="PdLd")
    axs.plot(UserData[0],UserData[4],drawstyle="steps",color='red',linewidth=0.8)

    axs.fill_between(UserData[0],UserData[5], step="pre", alpha=0.3,facecolor='darkgreen', label="PbAvSr")
    axs.plot(UserData[0],UserData[5],drawstyle="steps",color='darkgreen',linewidth=0.8)


    axs.fill_between(UserData[0],UserData[6],  step="pre", alpha=0.7,facecolor='orange', label="PbAvLd")
    axs.plot(UserData[0],UserData[6],drawstyle="steps",color='chocolate',linewidth=0.7)


    axs.fill_between(UserData[0],UserData[7], step="pre", alpha=0.35,facecolor='purple', label="PbRqLd")
    axs.plot(UserData[0],UserData[7],drawstyle="steps",color='purple',linewidth=0.8)


    Pgrd=np.add(UserData[1],UserData[2])
    axs.plot(UserData[0],Pgrd,drawstyle="steps",color='black',linewidth=1.5, label="Pgrd")
    axs.fill_between(UserData[0],Pgrd ,step="pre", alpha=0.3,facecolor='gray')

    axs.set_xlim(UserData[0][0],UserData[0][len(UserData[0])-1])

    axs.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
    axs.xaxis.set_major_locator(mdates.HourLocator(interval = 2))
    axs.yaxis.set_major_locator(MaxNLocator(integer=True))
    axs.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.2)


    axs.set_xlabel('Ura [h]',fontsize=11)
    axs.set_ylabel('P [kW]',fontsize=11)

    axs.xaxis.set_tick_params(labelsize=11)
    axs.yaxis.set_tick_params(labelsize=11)

    axs.set_ylim([1.1*PowerLow,1.1*PowerHigh])


def drawingPriceGraph(axs,UserData,PowerLow,PowerHigh):

    axs.plot(UserData[0],UserData[1],drawstyle="steps",color='blue',linewidth=2)

    axs.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
    axs.xaxis.set_major_locator(mdates.HourLocator(interval = 4))
    axs.yaxis.set_major_locator(MaxNLocator(integer=True))
    axs.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.2)

    axs.set_xlabel('Ura [h]',fontsize=11)
    axs.set_ylabel('¢/kWh',fontsize=11)

    axs.set_xlim(UserData[0][0],UserData[0][len(UserData[0])-1])
    axs.set_ylim([1.1*PowerLow,1.1*PowerHigh])
    

SelectedUser=[1,2,3,4]
FileDirecotory="./"
TestNumber=5
DateTime="01/02/2022 00:30"

drawingPowerGraph4Users(FileDirecotory, TestNumber, DateTime,SelectedUser)

 