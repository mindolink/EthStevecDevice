from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
import matplotlib.ticker as plticker
import matplotlib.ticker as ticker

import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import matplotlib.lines as mlines
from matplotlib.ticker import MultipleLocator
from matplotlib.transforms import BlendedGenericTransform

import numpy as np

def readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber):

    FolderPath=str(FileDirecotoryUserData)+'Test '+str(TestNumber)+' User '+str(UserNumber)+'.xlsx'

    wb = load_workbook(filename = FolderPath)
    xlsxPowerMeasurments = wb["PowerMeausurments"]

    row=36
    Pin=[0]*97
    Pout=[0]*97
    PdSr=[0]*97
    PdLd=[0]*97
    PbAvSr=[0]*97
    PbAvLd=[0]*97
    PbRqLd=[0]*97
    IntTime=[0]*97

    SOChsb=[0]*97
    SOCcar=[0]*97


    for q in range(97):
        Pout[q] -= xlsxPowerMeasurments["F"+str(row+q)].value
        Pin[q] = xlsxPowerMeasurments["E"+str(row+q)].value
        PdSr[q] -= xlsxPowerMeasurments["G"+str(row+q)].value
        PdLd[q] = xlsxPowerMeasurments["H"+str(row+q)].value
        PbAvSr[q] -= xlsxPowerMeasurments["I"+str(row+q)].value
        PbAvLd[q] = xlsxPowerMeasurments["J"+str(row+q)].value
        PbRqLd[q] = xlsxPowerMeasurments["K"+str(row+q)].value
        IntTime[q]= xlsxPowerMeasurments["B"+str(row+q)].value
        SOChsb[q] = xlsxPowerMeasurments["M"+str(row+q)].value
        SOCcar[q]= xlsxPowerMeasurments["O"+str(row+q)].value

        matrixUserData=[IntTime,Pout,Pin,PdSr,PdLd,PbAvSr,PbAvLd,PbRqLd,SOChsb,SOCcar]

    return(matrixUserData)


def readUserEnergyData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber):

    FolderPath=str(FileDirecotoryUserData)+'Test '+str(TestNumber)+' User '+str(UserNumber)+'.xlsx'

    wb = load_workbook(filename = FolderPath)
    xlsxPowerMeasurments = wb["EnergyMeausurments"]

    row=36
    Price=[0]*97
    Wout=[0]*97
    Win=[0]*97
    IntTime=[0]*97
    PriceKWH=[0]*97

    for q in range(97):

        Price[q]= xlsxPowerMeasurments["D"+str(row+q)].value
        Wout[q] = xlsxPowerMeasurments["F"+str(row+q)].value
        Win[q] = xlsxPowerMeasurments["E"+str(row+q)].value
        IntTime[q]= xlsxPowerMeasurments["B"+str(row+q)].value

        if Wout[q]!=0 or Win[q]!=0:
            PriceKWH[q]=(100*Price[q])/(Wout[q]+Win[q])/1000000000
        else:
            PriceKWH[q]=0

    print(PriceKWH)
    return(IntTime,PriceKWH)


def drawingPriceGraph4Users(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    UserData=[0]*4
    PriceLow=0
    PriceHigh=0

    for q in range (4):
        UserNumber=ArrUserNumber[q]
        UserData[q]=readUserEnergyData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)
        r=UserData[q]

        for y in range (1,2):
            for x in range (len(UserData[q][y])):
                PriceKWH=UserData[q][y][x]
                if PriceLow>PriceKWH:
                    PriceLow=PriceKWH

                if PriceKWH>PriceHigh:
                    PriceHigh=PriceKWH

    with plt.style.context(['science', 'grid']):

        fntSize=20

        fig = plt.figure(figsize=(14, 10),constrained_layout=True)
        gridsize = (33, 2)
        
        ax1=plt.subplot2grid(gridsize,(0, 0), colspan=1, rowspan=14)
        ax2=plt.subplot2grid(gridsize,(0, 1), colspan=1, rowspan=14)
        ax3=plt.subplot2grid(gridsize,(15, 0), colspan=1, rowspan=14)
        ax4=plt.subplot2grid(gridsize,(15, 1), colspan=1, rowspan=14)

        drawingPriceGraph(ax1,UserData[0],PriceLow,PriceHigh,fntSize)
        ax1.set_title('ODJEMALEC A',fontsize=fntSize)
        drawingPriceGraph(ax2,UserData[1],PriceLow,PriceHigh,fntSize)
        ax2.set_title('ODJEMALEC B',fontsize=fntSize)
        drawingPriceGraph(ax3,UserData[2],PriceLow,PriceHigh,fntSize)
        ax3.set_title('ODJEMALEC C',fontsize=fntSize)
        drawingPriceGraph(ax4,UserData[3],PriceLow,PriceHigh,fntSize)
        ax4.set_title('ODJEMALEC D',fontsize=fntSize)

        handles, labels = ax4.get_legend_handles_labels()
        leg=fig.legend( handles, labels,loc='upper center', bbox_to_anchor=(0.5,0.1),ncol=6,fontsize=fntSize+4,edgecolor="black",fancybox=False,handlelength=0.9,borderpad=0.3, frameon=False)
        leg.get_frame().set_linewidth(0.5)
        fig.set_constrained_layout_pads(hspace=0.4)

        plt.rcParams["font.family"] = "serif"
        plt.rcParams["font.serif"] = "Times New Roman"

        plt.savefig("Test "+str(TestNumber)+" Price.jpg", format="jpg")

        #plt.show()


def drawingPowerGraph4Users(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    UserData=[0]*4
    PowerLow=0
    PowerHigh=0

    for q in range (4):
        UserNumber=ArrUserNumber[q]
        UserData[q]=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)
        r=UserData[q]

        for y in range (1,8):
            for x in range (len(UserData[q][y])):
                Power=UserData[q][y][x]
                if PowerLow>Power:
                    PowerLow=Power

                if Power>PowerHigh:
                    PowerHigh=Power

    with plt.style.context(['science', 'grid']):
        
        fntSize=20

        fig = plt.figure(figsize=(14, 10),constrained_layout=True)
        gridsize = (33, 2)
        
        ax1=plt.subplot2grid(gridsize,(0, 0), colspan=1, rowspan=14)
        ax2=plt.subplot2grid(gridsize,(0, 1), colspan=1, rowspan=14)
        ax3=plt.subplot2grid(gridsize,(15, 0), colspan=1, rowspan=14)
        ax4=plt.subplot2grid(gridsize,(15, 1), colspan=1, rowspan=14)

        drawingPowerGraph(ax1,UserData[0],PowerLow,PowerHigh,fntSize)
        ax1.set_title('ODJEMALEC A',fontsize=fntSize)
        drawingPowerGraph(ax2,UserData[1],PowerLow,PowerHigh,fntSize)
        ax2.set_title('ODJEMALEC B',fontsize=fntSize)
        drawingPowerGraph(ax3,UserData[2],PowerLow,PowerHigh,fntSize)
        ax3.set_title('ODJEMALEC C',fontsize=fntSize)
        drawingPowerGraph(ax4,UserData[3],PowerLow,PowerHigh,fntSize)
        ax4.set_title('ODJEMALEC D',fontsize=fntSize)

        handles, labels = ax4.get_legend_handles_labels()
        leg=fig.legend( handles, labels,loc='upper center', bbox_to_anchor=(0.5,0.1),ncol=6,fontsize=fntSize+4,edgecolor="black",fancybox=False,handlelength=0.9,borderpad=0.3, frameon=False)
        
        leg.get_frame().set_linewidth(0.5)
        fig.set_constrained_layout_pads(hspace=0.4)

        plt.rcParams["font.family"] = "serif"
        plt.rcParams["font.serif"] = "Times New Roman"

        plt.savefig("Test "+str(TestNumber)+" Energy4User.jpg", format="jpg")

        #plt.show()

        
def drawingPowerSystemGraph(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    NumberOfUser=len(ArrUserNumber)
    UserData=[0]*NumberOfUser
    PowerLow=0
    PowerHigh=0
    
    for q in range (NumberOfUser):

        UserNumber=ArrUserNumber[q]
        UserData=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)

        if q==0:
            SystemData=UserData

        else:
            for y in range (1,8):
                for x in range (len(UserData[0])):
                    SystemData[y][x]+=UserData[y][x]

    for y in range (1,8):
        for x in range (len(SystemData[0])):
            Power=SystemData[y][x]
            if PowerLow>Power:
                PowerLow=Power
            if Power>PowerHigh:
                PowerHigh=Power
 

    with plt.style.context(['science', 'grid']):
    
        fntSize=24

        gridsize = (26, 24)
        fig = plt.figure(figsize=(14, 7),constrained_layout=True)
        ax=plt.subplot2grid(gridsize,(0, 1), colspan=22, rowspan=23)

        drawingPowerGraph(ax,SystemData,PowerLow,PowerHigh,fntSize)
        
        handles, labels = ax.get_legend_handles_labels()
        leg=fig.legend( handles, labels,loc='upper center', bbox_to_anchor=(0.5,0.11),ncol=6,fontsize=fntSize+6,edgecolor="black",fancybox=False,handlelength=0.9,borderpad=0.3, frameon=False)

        ax.set_ylabel('P [kW]',fontsize=fntSize)
        ax.set_xlabel('Ura [h]',fontsize=fntSize)

        ax.xaxis.set_tick_params(labelsize=fntSize)
        ax.yaxis.set_tick_params(labelsize=fntSize)
        
        leg.get_frame().set_linewidth(0.5)
        fig.set_constrained_layout_pads(hspace=0.4)

        plt.rcParams["font.family"] = "serif"
        plt.rcParams["font.serif"] = "Times New Roman"


        plt.savefig("Test "+str(TestNumber)+" EnergySystem.jpg", format="jpg")

        #plt.show()

def drawingPowerGraph(axs,UserData,PowerLow,PowerHigh,fntSize):


    axs.fill_between(UserData[0],UserData[3], step="pre", color="#2C71B0",alpha=0.20)
    axs.plot(UserData[0],UserData[3],drawstyle="steps", color="#2C71B0", alpha=1,linewidth=1.2,label=(r'$\mathrm{P_{dSr}}$'))

    axs.fill_between(UserData[0],UserData[4], step="pre", color="#FF4822",alpha=0.20)
    axs.plot(UserData[0],UserData[4],drawstyle="steps",color="#FF4822",alpha=1,linewidth=1.2,label=(r'$\mathrm{P_{dLd}}$'))

    axs.fill_between(UserData[0],UserData[5], step="pre", color="#14be53",alpha=0.20)
    axs.plot(UserData[0],UserData[5],drawstyle="steps", color="#14be53", alpha=1,linewidth=1.2,label=(r'$\mathrm{P_{bAvSr}}$'))

    axs.fill_between(UserData[0],UserData[6], step="pre", color="orange", alpha=0.20)
    axs.plot(UserData[0],UserData[6],drawstyle="steps", color="orange",alpha=1,linewidth=1.2,label=(r'$\mathrm{P_{bAvLd}}$'))

    axs.fill_between(UserData[0],UserData[7], step="pre",color="#dd37cc", alpha=0.20)
    axs.plot(UserData[0],UserData[7],drawstyle="steps", color="#dd37cc" , alpha=0.8,linewidth=1.2,label=(r'$\mathrm{P_{bRqLd}}$'))

    Pgrd=np.add(UserData[1],UserData[2])

    axs.fill_between(UserData[0],Pgrd, step="pre",color="#555555", alpha=0.18)
    axs.plot(UserData[0],Pgrd, drawstyle="steps",color="#555555", alpha=1,linewidth=1.2,label=(r'$\mathrm{P_{grd}}$'))
    
    axs.set_xlim(UserData[0][0],UserData[0][len(UserData[0])-1])

    axs.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
    axs.xaxis.set_major_locator(mdates.HourLocator(interval = 2))
    
    axs.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.15)

    axs.set_xlabel('Ura [h]',fontsize=fntSize)
    axs.set_ylabel('P [kW]',fontsize=fntSize)


    axs.xaxis.set_tick_params(labelsize=fntSize)
    axs.yaxis.set_tick_params(labelsize=fntSize)

    PowerHigh=1.1*math.ceil(PowerHigh)
    PowerLow=1.1*math.floor(PowerLow)

    axs.set_ylim([PowerLow,PowerHigh])

    axs.yaxis.set_major_locator(MaxNLocator(integer=True))
    axs.locator_params(axis="y", nbins=7)


def drawingPriceGraph(axs,UserData,PowerLow,PowerHigh,fntSize):

    axs.plot(UserData[0],[0]*len(UserData[0]),color="black",drawstyle="steps",linewidth=0.5)
    axs.plot(UserData[0],UserData[1],drawstyle="steps", color="blue" , alpha=0.8,linewidth=1.2,label='Cena AGR')
    

    axs.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
    axs.xaxis.set_major_locator(mdates.HourLocator(interval = 4))
    axs.yaxis.set_major_locator(MaxNLocator(integer=True))
    axs.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.2)

    axs.set_xlim(UserData[0][0],UserData[0][len(UserData[0])-1])

    axs.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
    axs.xaxis.set_major_locator(mdates.HourLocator(interval = 2))
    
    axs.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.15)

    axs.set_xlabel('Ura [h]',fontsize=fntSize)
    axs.set_ylabel('Cena EE [¢/kWh]',fontsize=fntSize)
        
    plt.rcParams["font.family"] = "serif"
    plt.rcParams["font.serif"] = "Times New Roman"

    axs.xaxis.set_tick_params(labelsize=fntSize)
    axs.yaxis.set_tick_params(labelsize=fntSize)


    PowerHigh=1.1*math.ceil(PowerHigh)
    PowerLow=1.1*math.floor(PowerLow)

    axs.set_ylim([PowerLow,PowerHigh])

    axs.yaxis.set_major_locator(MaxNLocator(integer=True))
    axs.locator_params(axis="y", nbins=7)

    axs.set_ylim(-23.5,23.5)

 
def drawingSOCGraaph(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    NumberOfUser=len(ArrUserNumber)

    with plt.style.context(['science', 'grid']):

        colorArray=["#2C71B0","#FF4822","#14be53","orange","#dd37cc","#555555"]
    
        fntSize=24

        gridsize = (26, 24)
        fig = plt.figure(figsize=(14, 7),constrained_layout=True)
        ax=plt.subplot2grid(gridsize,(0, 1), colspan=22, rowspan=23)

        i=0

        for q in range (NumberOfUser):

            Data=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,q+1)
            SOChsb=Data[8]
            SOCcar=Data[9]
            Time=Data[0]
            
            if SOChsb[0]!=None:

                lab="{$\mathrm{SBH_{"+str(chr(65+q))+"}"+"}"+"$}"
                
                ax.plot(Time,SOChsb, alpha=1,label=lab,color=colorArray[i],linewidth=1.2)
                i+=1
            if SOCcar[0]!=None:

                lab="{$\mathrm{SEV_{"+str(chr(65+q))+"}"+"}"+"$}"

                ax.plot(Time,SOCcar, alpha=1,label=lab,color=colorArray[i],linewidth=1.2)
                i+=1


        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
        ax.xaxis.set_major_locator(mdates.HourLocator(interval = 2))
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.25)


        leg=fig.legend(loc='upper center', bbox_to_anchor=(0.5,0.11),ncol=6,fontsize=fntSize,edgecolor="black",fancybox=False,handlelength=0.9,borderpad=0.3, frameon=False)
       
        ax.set_xlim(Time[0],Time[len(Time)-1])
        ax.set_ylim(0.1,100)
        ax.set_ylabel('Napolnjenost baterije [\%]',fontsize=fntSize)
        ax.set_xlabel('Ura [h]',fontsize=fntSize)

        ax.xaxis.set_tick_params(labelsize=fntSize)
        ax.yaxis.set_tick_params(labelsize=fntSize)

        leg.get_frame().set_linewidth(0.5)
        fig.set_constrained_layout_pads(hspace=0.4)

        plt.rcParams["font.family"] = "serif"
        plt.rcParams["font.serif"] = "Times New Roman"

        plt.savefig("Test "+str(TestNumber)+" SOC.jpg", format="jpg")
        #plt.show()



SelectedUser=[1,2,3,4]
FileDirecotory="./"
DateTime="01/02/2022 00:30"

for i in range (1,6):
    TestNumber=i
    drawingPowerGraph4Users(FileDirecotory, TestNumber, DateTime,SelectedUser)
    drawingPowerSystemGraph(FileDirecotory, TestNumber, DateTime,SelectedUser)
    drawingSOCGraaph(FileDirecotory, TestNumber, DateTime,SelectedUser)
    drawingPriceGraph4Users(FileDirecotory, TestNumber, DateTime,SelectedUser)