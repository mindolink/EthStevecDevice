from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
import matplotlib.ticker as plticker
import matplotlib.ticker as ticker

import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import matplotlib.lines as mlines
from matplotlib.ticker import MultipleLocator
from matplotlib.transforms import BlendedGenericTransform

import numpy as np

def readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber):

    FolderPath=str(FileDirecotoryUserData)+'Test '+str(TestNumber)+' User '+str(UserNumber)+'.xlsx'

    wb = load_workbook(filename = FolderPath)
    xlsxPowerMeasurments = wb["PowerMeausurments"]

    row=36
    Pin=[0]*97
    Pout=[0]*97
    PdSr=[0]*97
    PdLd=[0]*97
    PbAvSr=[0]*97
    PbAvLd=[0]*97
    PbRqLd=[0]*97
    IntTime=[0]*97

    SOChsb=[0]*97
    SOCcar=[0]*97


    for q in range(97):
        Pout[q] -= xlsxPowerMeasurments["F"+str(row+q)].value
        Pin[q] = xlsxPowerMeasurments["E"+str(row+q)].value
        PdSr[q] -= xlsxPowerMeasurments["G"+str(row+q)].value
        PdLd[q] = xlsxPowerMeasurments["H"+str(row+q)].value
        PbAvSr[q] -= xlsxPowerMeasurments["I"+str(row+q)].value
        PbAvLd[q] = xlsxPowerMeasurments["J"+str(row+q)].value
        PbRqLd[q] = xlsxPowerMeasurments["K"+str(row+q)].value
        IntTime[q]= xlsxPowerMeasurments["B"+str(row+q)].value
        SOChsb[q] = xlsxPowerMeasurments["M"+str(row+q)].value
        SOCcar[q]= xlsxPowerMeasurments["O"+str(row+q)].value

        matrixUserData=[IntTime,Pout,Pin,PdSr,PdLd,PbAvSr,PbAvLd,PbRqLd,SOChsb,SOCcar]

    return(matrixUserData)


def readUserEnergyData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber):

    FolderPath=str(FileDirecotoryUserData)+'Test '+str(TestNumber)+' User '+str(UserNumber)+'.xlsx'

    wb = load_workbook(filename = FolderPath)
    xlsxPowerMeasurments = wb["EnergyMeausurments"]

    row=36
    Price=[0]*97
    Wout=[0]*97
    Win=[0]*97
    IntTime=[0]*97
    PriceKWH=[0]*97

    for q in range(97):

        Price[q]= xlsxPowerMeasurments["D"+str(row+q)].value
        Wout[q] = xlsxPowerMeasurments["F"+str(row+q)].value
        Win[q] = xlsxPowerMeasurments["E"+str(row+q)].value
        IntTime[q]= xlsxPowerMeasurments["B"+str(row+q)].value

        if Wout[q]!=0 or Win[q]!=0:
            PriceKWH[q]=(100*Price[q])/(Wout[q]+Win[q])/1000000000
        else:
            PriceKWH[q]=0

    print(PriceKWH)
    return(IntTime,PriceKWH)


def drawingPriceGraph4Users(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    UserData=[0]*4
    PriceLow=0
    PriceHigh=0

    for q in range (4):
        UserNumber=ArrUserNumber[q]
        UserData[q]=readUserEnergyData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)
        r=UserData[q]

        for y in range (1,2):
            for x in range (len(UserData[q][y])):
                PriceKWH=UserData[q][y][x]
                if PriceLow>PriceKWH:
                    PriceLow=PriceKWH

                if PriceKWH>PriceHigh:
                    PriceHigh=PriceKWH

    with plt.style.context(['science', 'grid']):

        fig, axs = plt.subplots(2, 2,figsize=(13.6,8),constrained_layout=True)

        for q in range (4):

            UserLabel=chr(64+(ArrUserNumber[q]))

            if q<2:
                drawingPriceGraph(axs[0][q],UserData[q],PriceLow,PriceHigh)
                axs[0][q].set_title('UPORABNIK '+str(UserLabel),fontsize=18)
            
            else:
                drawingPriceGraph(axs[1][q-2],UserData[q],PriceLow,PriceHigh)
                axs[1][q-2].set_title('UPORABNIK '+str(UserLabel),fontsize=18)


            if q==3:
                axs[1][q-2].legend(bbox_to_anchor=(1.04,0), loc="lower left",ncol=1,fontsize=18)


            #plt.rcParams["font.family"] = "serif"
            #plt.rcParams["font.serif"] = "Times New Roman"


        fig.set_constrained_layout_pads(hspace=0.07)
        #plt.show()
        plt.savefig("Test "+str(TestNumber)+" PRC.jpg", format="jpg")


def drawingPowerGraph4Users(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    UserData=[0]*4
    PowerLow=0
    PowerHigh=0

    for q in range (4):
        UserNumber=ArrUserNumber[q]
        UserData[q]=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)
        r=UserData[q]

        for y in range (1,8):
            for x in range (len(UserData[q][y])):
                Power=UserData[q][y][x]
                if PowerLow>Power:
                    PowerLow=Power

                if Power>PowerHigh:
                    PowerHigh=Power

    with plt.style.context(['science', 'grid']):

        gridsize = (26, 2)
        fig = plt.figure(figsize=(14, 10),constrained_layout=True)


        ax1=plt.subplot2grid(gridsize,(0, 0), colspan=1, rowspan=11)
        ax2=plt.subplot2grid(gridsize,(0, 1), colspan=1, rowspan=11)
        ax3=plt.subplot2grid(gridsize,(12, 0), colspan=1, rowspan=11)
        ax4=plt.subplot2grid(gridsize,(12, 1), colspan=1, rowspan=11)

        drawingPowerGraph(ax1,UserData[0],PowerLow,PowerHigh)
        ax1.set_title('ODJEMALEC A',fontsize=18)

        drawingPowerGraph(ax2,UserData[1],PowerLow,PowerHigh)
        ax2.set_title('ODJEMALEC B',fontsize=18)

        drawingPowerGraph(ax3,UserData[2],PowerLow,PowerHigh)
        ax3.set_title('ODJEMALEC C',fontsize=18)

        drawingPowerGraph(ax4,UserData[3],PowerLow,PowerHigh)
        ax4.set_title('ODJEMALEC D',fontsize=18)

        handles, labels = ax4.get_legend_handles_labels()
        fig.legend( handles, labels,loc='upper center', bbox_to_anchor=(0.5,0.07),ncol=6,fontsize=18)

        fig.set_constrained_layout_pads(hspace=0.4)

        plt.rcParams["font.family"] = "serif"
        plt.rcParams["font.serif"] = "Times New Roman"

        
        #plt.savefig("2222test.svg", format="svg")
        #plt.savefig("Test "+str(TestNumber)+" POW.jpg", format="jpg")

        plt.show()

        
def drawingPowerSystemGraph(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    NumberOfUser=len(ArrUserNumber)
    UserData=[0]*NumberOfUser
    PowerLow=0
    PowerHigh=0
    
    for q in range (NumberOfUser):

        UserNumber=ArrUserNumber[q]
        UserData=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,UserNumber)

        if q==0:
            SystemData=UserData

        else:
            for y in range (1,8):
                for x in range (len(UserData[0])):
                    SystemData[y][x]+=UserData[y][x]


    for y in range (1,8):
        for x in range (len(SystemData[0])):
            Power=SystemData[y][x]
            if PowerLow>Power:
                PowerLow=Power
            if Power>PowerHigh:
                PowerHigh=Power
 

    with plt.style.context(['science', 'grid']):
    
        fig, ax = plt.subplots(figsize=(12,8),constrained_layout=True)
        drawingPowerGraph(ax,SystemData,PowerLow,PowerHigh)
        ax.legend(bbox_to_anchor=(1.04,0), loc="lower left",ncol=1,fontsize=18)
        #plt.show()

        plt.savefig("Test "+str(TestNumber)+" SYS.jpg", format="jpg")

def drawingPowerGraph(axs,UserData,PowerLow,PowerHigh):


    axs.fill_between(UserData[0],UserData[3], step="pre", color="#2C71B0",alpha=0.15)
    axs.plot(UserData[0],UserData[3],drawstyle="steps", color="#2C71B0", alpha=1,label="PdSr")

    axs.fill_between(UserData[0],UserData[4], step="pre", color="#FF4822",alpha=0.15)
    axs.plot(UserData[0],UserData[4],drawstyle="steps",color="#FF4822",alpha=1,label=(u"P$\\{dLd}$"))

    axs.fill_between(UserData[0],UserData[5], step="pre", color="#14be53",alpha=0.15)
    axs.plot(UserData[0],UserData[5],drawstyle="steps", color="#14be53", alpha=1,label="PbAvSr")

    axs.fill_between(UserData[0],UserData[6], step="pre", color="orange", alpha=0.20)
    axs.plot(UserData[0],UserData[6],drawstyle="steps", color="orange",alpha=1,label="PbAvLd")

    axs.fill_between(UserData[0],UserData[7], step="pre",color="#AA64CC", alpha=0.15)
    axs.plot(UserData[0],UserData[7],drawstyle="steps", color="#AA64CC" , alpha=0.8,label="PbRqLd")


    Pgrd=np.add(UserData[1],UserData[2])

    axs.fill_between(UserData[0],Pgrd, step="pre",color="#555555", alpha=0.18)
    axs.plot(UserData[0],Pgrd, drawstyle="steps",color="#555555", alpha=1,label="Pgrd")
    

    axs.set_xlim(UserData[0][0],UserData[0][len(UserData[0])-1])

    axs.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
    axs.xaxis.set_major_locator(mdates.HourLocator(interval = 2))
    
    axs.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.13)

    axs.set_xlabel('Ura [h]',fontsize=18)
    axs.set_ylabel('P [kW]',fontsize=18)


    axs.xaxis.set_tick_params(labelsize=18)
    axs.yaxis.set_tick_params(labelsize=18)

    vzr=6

    PowerHigh=1.1*math.ceil(PowerHigh)
    PowerLow=1.1*math.floor(PowerLow)

    
    axs.set_ylim([PowerLow,PowerHigh])

    axs.yaxis.set_major_locator(MaxNLocator(integer=True))
    axs.locator_params(axis="y", nbins=7)

  

def drawingPriceGraph(axs,UserData,PowerLow,PowerHigh):

    axs.plot(UserData[0],[0]*len(UserData[0]),color="black",drawstyle="steps",linewidth=0.5)
    axs.plot(UserData[0],UserData[1],drawstyle="steps",linewidth=1,label="Cena agregatorja")
    

    axs.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
    axs.xaxis.set_major_locator(mdates.HourLocator(interval = 4))
    axs.yaxis.set_major_locator(MaxNLocator(integer=True))
    axs.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.2)

    axs.set_xlabel('Ura [h]',fontsize=18)
    axs.set_ylabel('¢/kWh',fontsize=18)

    axs.set_xlim(UserData[0][0],UserData[0][len(UserData[0])-1])

    axs.xaxis.set_tick_params(labelsize=18)
    axs.yaxis.set_tick_params(labelsize=18)
    
    
    axs.set_ylim(-24,24)

    

def drawingSOCGraaph(FileDirecotoryUserData, TestNumber, DatetTimeTest, ArrUserNumber):

    NumberOfUser=len(ArrUserNumber)

    with plt.style.context(['science', 'grid']):

        fig, ax = plt.subplots(figsize=(12,8),constrained_layout=True)

        for q in range (NumberOfUser):
            Data=readUserPowerData(FileDirecotoryUserData, TestNumber, DatetTimeTest,q+1)
            SOChsb=Data[8]
            SOCcar=Data[9]
            Time=Data[0]

            if SOChsb[0]!=None:
                ax.plot(Time,SOChsb, alpha=1,label="SOChsb"+str(chr(65+q)))

            if SOCcar[0]!=None:
                ax.plot(Time,SOCcar, alpha=1,label="SOCcar"+str(chr(65+q)))


        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H'))
        ax.xaxis.set_major_locator(mdates.HourLocator(interval = 4))
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.grid(b=True, which='major', color='#444444', linestyle='-', alpha=0.2)

        ax.set_xlabel('Ura [h]',fontsize=18)
        ax.set_ylabel('SOC [%]',fontsize=18)

        ax.set_xlim(Time[0],Time[len(Time)-1])
        ax.set_ylim(0,100)

        ax.xaxis.set_tick_params(labelsize=18)
        ax.yaxis.set_tick_params(labelsize=18)
        
        ax.legend(bbox_to_anchor=(1.04,0), loc="lower left",ncol=1,fontsize=18)
        #plt.show()

        plt.savefig("Test "+str(TestNumber)+" SOC.jpg", format="jpg")







SelectedUser=[1,2,3,4]
FileDirecotory="./"
TestNumber=5
DateTime="01/02/2022 00:30"

drawingPowerGraph4Users(FileDirecotory, TestNumber, DateTime,SelectedUser)

#drawingPowerSystemGraph(FileDirecotory, TestNumber, DateTime,SelectedUser)
#drawingSOCGraaph(FileDirecotory, TestNumber, DateTime,SelectedUser)
#drawingPriceGraph4Users(FileDirecotory, TestNumber, DateTime,SelectedUser)