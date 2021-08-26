from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np

class carBattery(object):

    def __init__(self,UserNumber,NumberOfCars,FileDirectory):

        wb = load_workbook(filename = FileDirectory)
        xlsxBatteryProperties = wb['userCarProperties']
        row=UserNumber+3

        row=UserNumber+3
        #Init user stationary electric battery
        k=1000
        self.CarNum=NumberOfCars

        colume= get_column_letter(4+7*NumberOfCars)
        self.Wb=(xlsxBatteryProperties[str(colume)+str(row)].value)*k
        colume= get_column_letter(5+7*NumberOfCars)
        self.PbCh=(xlsxBatteryProperties[str(colume)+str(row)].value)*k
        colume= get_column_letter(6+7*NumberOfCars)
        self.PbDh=(xlsxBatteryProperties[str(colume)+str(row)].value)*k
        colume= get_column_letter(7+7*NumberOfCars)
        self.EffCh=(xlsxBatteryProperties[str(colume)+str(row)].value)/100
        colume= get_column_letter(8+7*NumberOfCars)
        self.EffDh=(xlsxBatteryProperties[str(colume)+str(row)].value)/100
        colume= get_column_letter(9+7*NumberOfCars)
        self.SOCmax=(xlsxBatteryProperties[str(colume)+str(row)].value)/100
        colume= get_column_letter(10+7*NumberOfCars)
        self.SOCmin=(xlsxBatteryProperties[str(colume)+str(row)].value)/100

        wb.close()

        print ("Propertise of Elektric car "+str(NumberOfCars)+":")
        print ('Wb:'+str(self.Wb/k)+'kWh  '+' PbCh:'+str(self.PbCh/k)+'kW  '+' PbDh:'+str(self.PbDh/k)+'kW  '+' ηCh:'+str(self.EffCh*100)+'%  '+
                ' ηDh:'+str(self.EffDh*100)+'%  '+' SOCmin:'+str(self.SOCmin*100)+' %  '+' SOCmax:'+str(self.SOCmax*100)+' %  ')

     
        self.BatOn="OFF"    #Start is power off
        self.BatSet=0       #Battery settings
        self.SOC=0  

        self.Pavg=0
        self.Wavg=0         #Avarage power 
        self.Pcur=0         #Current power
        self.Wcur=0
        self.Flg=0          #Flag for reset loop meausment average power
        self.Wsum=0
        self.Psum=0

        self.WeekNumber=0
        self.Hour=0
        self.TarNum=0       #Price tarif number
        self.OffOn=True        #Flag for init when car is connect to grid

        self.PbAvSr=0
        self.PbAvLd=0
        self.PbRqLd=0  


    def processingBatterySetting(self,BatOn,BatSet,SOCstart,WeekNumber,Hour,TarNum,HomNedEne,SysNedEne):
    
        self.BatOn=BatOn
        self.BatSet=BatSet
        self.TarNum=TarNum
        self.HomNedEne=HomNedEne
        self.SysNedEne=SysNedEne
        self.Hour=Hour

        if (self.BatOn=="ON"):

            if (self.OffOn==True):
                self.SOC=SOCstart/100
                self.OffOn=False

            if (self.BatSet==1):
                self.batteryFunctionSettings1()
            elif (self.BatSet==2):
                self.batteryFunctionSettings2()
            elif (self.BatSet==3):
                self.batteryFunctionSettings3()
            else:
                self.PbAvSr=0
                self.PbAvLd=0
                self.PbRqLd=0
             
        else:
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=0
            self.SOC=0
            self.OffOn=True
            
        #Display battery settings

        k=1000 #conversion factor from W to kW
        pavsr=("%.2f" % (self.PbAvSr/k))
        pavld=("%.2f" % (self.PbAvLd/k))
        prqld=("%.2f" % (self.PbRqLd/k))

        print ("Car"+str(self.CarNum+1)+" battery settings: PbAvSr:"+str(pavsr)+"kW   PbAvLd:"+str(pavld)+"kW  PbRqLd:"+str(prqld)+"kW")

    def getRequiredPower(self):
        return ([self.PbAvSr,self.PbAvLd,self.PbRqLd])

    def batteryFunctionSettings1(self):

        if (self.TarNum==3 and self.SOCmin<self.SOC and (self.HomNedEne==True and self.SysNedEne==True)):
            self.PbAvSr=self.PbDh
            self.PbAvLd=0
            self.PbRqLd=0

        elif (self.SOC<self.SOCmax):
            self.PbAvSr=0
            self.PbAvLd=self.PbCh
            self.PbRqLd=0

        else:
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=0  
        
        
    def batteryFunctionSettings2(self):

        if (self.TarNum==1 and self.SOCmax>self.SOC and (self.Hour<6 or self.Hour>21)):
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=self.PbCh

        elif (self.SOC<self.SOCmax):
            self.PbAvSr=0
            self.PbAvLd=self.PbCh
            self.PbRqLd=0  
        else:
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=0  

    def batteryFunctionSettings3(self):

        if (self.SOC<self.SOCmax):
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=self.PbCh       
        else:
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=0

    def setBatteryPower(self,puP):

        self.Pcur=-puP[2]*self.PbAvSr+puP[3]*self.PbAvLd+puP[4]*self.PbRqLd
        self.Psum+=self.Pcur

    def updateBatteryValues(self,dt):

        if (self.Pcur>0):
            self.Wcur=(self.Pcur*self.EffCh*dt)/3600
            self.SOC=(((self.SOC*self.Wb)+self.Wcur)/self.Wb)
            self.Wsum+=self.Pcur*dt
            

        else:
            self.Wcur=(self.Pcur*self.EffDh*dt)/3600
            self.SOC=(((self.SOC*self.Wb)+self.Wcur)/self.Wb)
            self.Wsum+=self.Pcur*dt
            
        
    def getBatteryInfo(self,Flg):

        self.Pavg=self.Psum/Flg
        self.Wavg=self.Wsum

        self.Wsum=0
        self.Psum=0

        return ([self.Pavg,self.Wavg,self.SOC])