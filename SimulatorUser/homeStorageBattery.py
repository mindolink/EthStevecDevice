from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np

class homeStorageBattery(object):

    def __init__(self,UserNumber,FileDirectory):

        wb = load_workbook(filename = FileDirectory)
        xlsxBatteryProperties = wb['userHomeProperties']
        row=UserNumber+3

        #Init user stationary electric battery
        k=1000


        #Init user stationary electric battery
        status=(xlsxBatteryProperties['D'+str(row)].value)
        if (status=="ON"):
            self.BatOn=True
            self.Wb=(xlsxBatteryProperties["E"+str(row)].value)*k
            self.PbCh=(xlsxBatteryProperties["F"+str(row)].value)*k
            self.PbDh=(xlsxBatteryProperties["G"+str(row)].value)*k
            self.EffCh=(xlsxBatteryProperties["H"+str(row)].value)/100
            self.EffDh=(xlsxBatteryProperties["I"+str(row)].value)/100
            self.SOCmax=(xlsxBatteryProperties["J"+str(row)].value)/100
            self.SOCmin=(xlsxBatteryProperties["K"+str(row)].value)/100

            print ("Propertise of Home Storage Battery:")
        
            print ('Wb:'+str(self.Wb/k)+'kWh  '+' PbCh:'+str(self.PbCh/k)+'kW  '+' PbDh:'+str(self.PbDh/k)+'kW  '+' ηCh:'+str(self.EffCh*100)+'%  '+
            ' ηDh:'+str(self.EffDh*100)+'%  '+' SOCmin:'+str(self.SOCmin*100)+' %  '+' SOCmax:'+str(self.SOCmax*100)+' %  ')

        else:
            self.BatOn=False
            self.Wb=0
            self.PbCh=0
            self.PbDh=0
            self.EffCh=0
            self.EffDh=0
            self.SOCmax=0
            self.SOCmin=0
            print ("User don't have Home Storage Battery!")

        wb.close()
 
        self.SOC=0.3

        self.Pavg=0
        self.Wavg=0         #Avarage power 
        self.Pcur=0         #Current power
        self.Wcur=0
        self.Flg=0          #Flag for reset loop meausment average power
        self.Wsum=0
        self.Psum=0

        self.WeekNumber=0
        self.Hour=0
        self.TarNum=0       #Price tarif number
        self.OffOn=True        #Flag for init when car is connect to grid

        self.PbAvSr=0
        self.PbAvLd=0
        self.PbRqLd=0  


    def processBatterySetting(self,SOCsmart,WeekNumber,Hour,TarNum,TarInt,HomNedEne,SysNedEne):

        if self.BatOn==True:
            
            self.WeekNumber=WeekNumber
            self.Hour=Hour
            self.SOCsmart=SOCsmart/100
            self.HomNedEne=HomNedEne
            self.SysNedEne=SysNedEne
            self.TarNum=TarNum
            self.TarInt=TarInt

            if (self.TarNum==1):
                self.settingsTariff1()
            elif (self.TarNum==2):
                self.settingsTariff2()
            elif (self.TarNum==3):
                self.settingsTariff3()
            else:
                    self.PbAvSr=0
                    self.PbAvLd=0
                    self.PbRqLd=0
        else:
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=0
            
        #Display battery settings

        k=1000 #Conversion factor from W to kW
        pavsr=("%.2f" % (self.PbAvSr/k))
        pavld=("%.2f" % (self.PbAvLd/k))
        prqld=("%.2f" % (self.PbRqLd/k))
               
        print ("Home battery settings: PbAvSr:"+str(pavsr)+"kW   PbAvLd:"+str(pavld)+"kW  PbRqLd:"+str(prqld)+"kW")


    def getRequiredPower(self):

        self.setPb=[self.PbAvSr,self.PbAvLd,self.PbRqLd]

        return(self.setPb)

    def settingsTariff1(self):
        if (self.WeekNumber<6):
            if (self.SOC<self.SOCsmart and (self.Hour<6 or 21<self.Hour)):
                dtWb=(self.SOCsmart-self.SOC)*self.Wb
                P=dtWb/(self.TarInt*0.9)
                if P<self.PbCh:
                    self.PbAvSr=0
                    self.PbAvLd=self.PbCh-P
                    self.PbRqLd=P
                else:
                    self.PbAvSr=0
                    self.PbAvLd=self.PbCh
                    self.PbRqLd=0

            elif (self.SOC<self.SOCmax):
                self.PbAvSr=0
                self.PbAvLd=self.PbCh
                self.PbRqLd=0
            else:
                self.PbAvSr=0
                self.PbAvLd=0
                self.PbRqLd=0
        else:
            if (self.Haour>15 and SOC>SOCmin and self.HomNedEne==True and self.SysNedEne==True):
                self.PbAvSr=self.PbDh
                self.PbAvLd=0
                self.PbRqLd=0
            elif (self.SOC<self.SOCmax):
                self.PbAvSr=0
                self.PbAvLd=self.PbCh
                self.PbRqLd=0
            else:
                self.PbAvSr=0
                self.PbAvLd=0
                self.PbRqLd=0

    def settingsTariff2(self):

        if (self.SOC<self.SOCmax):
            self.PbAvSr=0
            self.PbAvLd=self.PbCh
            self.PbRqLd=0

        else:
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=0

    def settingsTariff3(self):

        if (self.SOC>self.SOCmin and self.HomNedEne==True and self.SysNedEne==True):
            self.PbAvSr=self.PbDh
            self.PbAvLd=0
            self.PbRqLd=0
        
        
        elif (self.SOC<self.SOCmax):
            self.PbAvSr=0
            self.PbAvLd=self.PbCh
            self.PbRqLd=0

        else:
            self.PbAvSr=0
            self.PbAvLd=0
            self.PbRqLd=0
    

    def setBatteryPower(self,puP):

        self.Pcur=-puP[2]*self.PbAvSr+puP[3]*self.PbAvLd+puP[4]*self.PbRqLd
        self.Psum+=self.Pcur
            

    def updateBatteryValues(self,dt):

        if (self.Pcur>0):
            self.Wcur=(self.Pcur*self.EffCh*dt)/3600
            self.SOC=(((self.SOC*self.Wb)+self.Wcur)/self.Wb)
            self.Wsum+=self.Pcur*dt
        else:
            self.Wcur=(self.Pcur*self.EffDh*dt)/3600
            self.SOC=(((self.SOC*self.Wb)+self.Wcur)/self.Wb)
            self.Wsum+=self.Pcur*dt
        
    def getBatteryInfo(self,Flg):

        self.Pavg=self.Psum/Flg
        self.Wavg=self.Wsum

        self.Wsum=0
        self.Psum=0

        return ([self.Pavg,self.Wavg,self.SOC])