import time
import carBattery,homeStorageBattery,batteryManegmentSystem,linkEthNetwork,savingMeasurements,address
import numpy as np

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


k=1000
nano=1000000000
a=0
UserAccount=5
Http='http://localhost:8545'
AddressSCB=address.SCB
AddressSCC=address.SCC
PathUserInfo='./ImportData/userInfo.xlsx'
PathUserSchedule='./ImportData/userSchedule.xlsx'
PathAbiSCC='./SmartConcract/abiSystemControlingConcract.json'
PathAbiSCB='./SmartConcract/abiElectricityBillingConcract.json'

dt=30
t=1
DayName=[0,'MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSTDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']

Sec=0
Min=0
Hour=0
Day=0

StrFlg=True
SecFlg=True
MinFlg=True
AvgFlg=True

TarInt=0
SystemRun=False
SystemNeedEnergy=False
EnergyMeter=0

ReqArrPower=[0]*5
SndArrPower=[0]*5
GetArrPower=[0]*5
ActArrPower=[0]*5
SumArrGrdEnergy=[0]*5
SumArrGrdPower=[0]*5
SumArrTotEnergy=[0]*5
SumArrTotPower=[0]*5

#-------------Init LEP module-------------
ethReg=linkEthNetwork.systemControling(AddressSCC,PathAbiSCC,Http)
ethBil=linkEthNetwork.electricityBilling(AddressSCB,PathAbiSCB,Http)

#-------------Auto registration in the register of Smart Concract SCC and SCB-------------
if ethReg.getUserIndex()==0:
    ethReg.autoRegistrationNewUser()
    ethBil.autoRegistrationNewUser()
    time.sleep(5)
                    
    
UserIndex=ethReg.getUserIndex()
TestNumber=ethReg.getTestNumber()
BlockNumber=ethReg.getBlock()

print(UserIndex)

#-------------Init BMS module-------------
bms=batteryManegmentSystem.batteryManegmentSystem()


#-------------Init propertise HSB module-------------
hsb=homeStorageBattery.homeStorageBattery(UserIndex,PathUserInfo)


#-------------Init propertise CAR module-------------

wbInfo = load_workbook(filename = PathUserInfo)
xlsUserCars= wbInfo['userCarProperties']
NumberOfCars=int(xlsUserCars["C"+str(UserIndex+3)].value)

if NumberOfCars>0:
    car=[0]*NumberOfCars
    for q in range (NumberOfCars):
        car[q]=carBattery.carBattery(UserIndex,q,PathUserInfo)


#-------------Init propertise EXD module-------------

sm=savingMeasurements.savingMeasurements(UserIndex,TestNumber,NumberOfCars)



#-------------Read the initial values-------------


wbSchedule = load_workbook(filename = PathUserSchedule)
xlsxUserSchedule = wbSchedule["User "+str(UserIndex)]
StrDay=xlsxUserSchedule["B4"].value

StrRow=4

r=0

#-------------Loop program-------------

while r<23:
    
    if SystemRun==True:
        
        SystemRun=ethReg.getSystemRun()
        SystemNeedEnergy=ethReg.getSystemNeedEnergy()

        if (StrFlg==False):

            StrTime=time.time_ns()
            Row=StrRow
            NumSec=1

            RowFlg=True
            SecFlg=True
            MinFlg=True
            AvgFlg=True
            StrFlg=True

#-------------Read date and time-------------
        print(xlsxUserSchedule["B"+str(Row)].value)
        DateTime = xlsxUserSchedule["B"+str(Row)].value
        WeekNumber=datetime.date(DateTime).weekday()+1

        print("DATE: "+str(DayName[WeekNumber])+" "+str(DateTime.strftime("%d/%m/%Y")))
        print("TIME: "+str(Hour)+":"+str(Min)+":"+str(Sec))

        if Min==0:
            DateTimeStr=(DateTime.strftime("%d/%m/%Y %H"))+":0"+str(Min)
        else:
            DateTimeStr=(DateTime.strftime("%d/%m/%Y %H"))+":"+str(Min)

        
#-------------Time tariff interval-------------

        TarNum=xlsxUserSchedule["C"+str(Row)].value
        TarInt=0

        for q in range (24):
            RowInt=Row+q
            TarLop=xlsxUserSchedule["C"+str(RowInt)].value

            if (TarLop==TarNum):
                TarInt+=1
            else:
                break

#-------------Read power pasive production and consumption-------------           

        ReqPdSr=(xlsxUserSchedule["D"+str(Row)].value)
        ReqPdLd=(xlsxUserSchedule["E"+str(Row)].value)

        if ReqPdSr>ReqPdLd:
            HomNedEne=False
        else:
            HomNedEne=True

#-------------------Read the settings HSB module-------------      
        print ("")
        print("BATTERY SETINGS:")

        SOCsmart=xlsxUserSchedule["F"+str(Row)].value
        hsb.processBatterySetting(SOCsmart,WeekNumber,Hour,TarNum,TarInt,HomNedEne,SystemNeedEnergy)
        ReqPhsb=hsb.getRequiredPower()

#-------------------Read the settings CAR module-------------      

        ReqPcar=[0]*3

        if NumberOfCars>0:

            for q in range (NumberOfCars):

                BatOn=0
                BatSet=0
                SOCstart=0

                try:
                    Colume= get_column_letter(7+3*q)
                    BatOn=xlsxUserSchedule[str(Colume)+str(Row)].value

                    Colume= get_column_letter(8+3*q)
                    BatSet=xlsxUserSchedule[str(Colume)+str(Row)].value

                    Colume= get_column_letter(9+3*q)
                    SOCstart=xlsxUserSchedule[str(Colume)+str(Row)].value

                except:
                    BatSet=0
                    SOCstart=0

                car[q].processingBatterySetting(BatOn,BatSet,SOCstart,WeekNumber,Hour,TarNum,HomNedEne,SystemNeedEnergy)
                ReqOnePcar=car[q].getRequiredPower()
                ReqPcar=np.add(ReqPcar,ReqOnePcar)

#-------------------Total power production and consumption-------------      
        
        ReqPbat=np.add(ReqPcar,ReqPhsb)

        ReqArrPower[0]=ReqPdSr
        ReqArrPower[1]=ReqPdLd
        ReqArrPower[2]=ReqPbat[0]
        ReqArrPower[3]=ReqPbat[1]
        ReqArrPower[4]=ReqPbat[2]

        print ("")
        print ("REQUAST POWERS:")
        print("PdSr:"+str(round(ReqArrPower[0]/k,2))+"kW  PdLd:"+str(round(ReqArrPower[1]/k,2))+"kW  PbAvSr:"
        +str(round(ReqArrPower[2]/k,2))+"kW  PbAvLd:"+str(round(ReqArrPower[3]/k,2))+"kW  PbRqLd:"
        +str(round(ReqArrPower[4]/k,2))+"kW")


#-------------------Control BMS system limitations-------------    

        bms.processAllParametersAndRestrictions(ReqArrPower,GetArrPower)
        SndReqPower=bms.inputPowerDataInfoForConcract()


#-------------------Sending data power requirements and wishes in ETH-------------    

        if ethReg.checkBlock():
            ethReg.setUserDataPower(SndReqPower)
        

#-------------------Set system power-------------    

        SystemNeedEnergy=ethReg.getSystemNeedEnergy()
        GetArrPower=ethReg.getUserDataPower()
        bms.processAllParametersAndRestrictions(ReqArrPower, GetArrPower)

        ActArrTotPower=bms.actualTotalPower()
        ActArrGrdPower=bms.actualPowerFromOrToGrid()

        SumArrTotPower=np.add(SumArrTotPower,ActArrTotPower)
        SumArrGrdPower=np.add(SumArrGrdPower,ActArrGrdPower)

        puActArrPower=bms.peerUnitRequestedPower()


        print ("")
        print ("ACTUAL POWERS IN SEGMENT:")
        print("PdSr:"+str(round(ActArrTotPower[0]/k,2))+"kW  PdLd:"+str(round(ActArrTotPower[1]/k,2))+"kW  PbAvSr:"
        +str(round(ActArrTotPower[2]/k,2))+"kW  PbAvLd:"+str(round(ActArrTotPower[3]/k,2))+"kW  PbRqLd:"
        +str(round(ActArrTotPower[4]/k,2))+"kW")
        print ("")


#-------------------Set device power-------------   

        hsb.setBatteryPower(puActArrPower)

        if NumberOfCars>0:
            for q in range (NumberOfCars):
                car[q].setBatteryPower(puActArrPower)


#-------------------Sending data energy production and consumption in ETH-------------  

        if ((Min==0 or Min==15 or Min==30 or Min==45) and Sec==0):

            Row=int(StrRow+(Min/15)+4*((24*Day)+Hour))
            print ("LEEEEP")
            xlsxSystemTarifPrices = wbInfo["systemTariffPrices"]
            TarNumPre=xlsxUserSchedule["C"+str(Row)].value

            PriceBuy=xlsxSystemTarifPrices["C"+str(TarNumPre+2)].value
            PriceSell=xlsxSystemTarifPrices["D"+str(TarNumPre+2)].value
            
            ethBil.modifaySystemTarifPrice(int(TarNumPre),int(PriceBuy), int(PriceSell))
            ethBil.setUserDataEnergy([int(SumArrGrdEnergy[0]),int(SumArrGrdEnergy[1]),int(SumArrGrdEnergy[2]),int(SumArrGrdEnergy[3]),int(SumArrGrdEnergy[4])])
            
       

#-------------------Storing power and energy measurement data------------- 

            AvgArrTotPower=np.divide(SumArrTotPower,AvgFlg)

            AvgPin=0
            AvgPout=0
            SumEin=0
            SumEout=0

            for q in range(5):
                if q==0 or q==2:
                    AvgPout+=SumArrGrdPower[q]/AvgFlg
                    SumEout+=SumArrGrdEnergy[q]

                else:
                    AvgPin+=SumArrGrdPower[q]/AvgFlg
                    SumEin+=SumArrGrdEnergy[q]

            sm.safeBasicMeasurements(DateTimeStr, AvgPout, AvgPin, AvgArrTotPower, SumEin, SumEout, SumArrTotEnergy)


            if (hsb.BatOn==True):
                InfoBat=hsb.getBatteryInfo(AvgFlg)
                sm.safeHomeBatteryMeasurements(InfoBat)

            if (NumberOfCars>0):
                for q in range (NumberOfCars):
                    InfoBat=car[q].getBatteryInfo(AvgFlg)
                    sm.safeCarBatteryMeasurements(q, InfoBat)

            SumArrGrdEnergy=[0]*5
            SumArrTotEnergy=[0]*5
            SumArrGrdPower=[0]*5
            SumArrTotPower=[0]*5

            AvgFlg=0
            
#-------------------Billing for energy production and consumption in ETH-------------------

        if ((Min==5 or Min==20 or Min==35 or Min==50) and Sec==0):

            ethBil.processingBillingForEnergy()

#-------------------Storing price and wallet data-------------------

        if ((Min==10 or Min==25 or Min==40 or Min==55) and Sec==0):

            MonayWalletCent=ethBil.getUserWalletInCent()
            PriceForEnergyCent=ethBil.getUserFinalEnergyPriceInCent()
            sm.safeCashBalance(MonayWalletCent, PriceForEnergyCent)


#-------------------Internal time-------------------
        
        AvgFlg+=1
        NumSec+=1
        Sec+=dt

        if Sec>=60:
            Min+=1
            Sec=0

        if Min>=60:
            Hour+=1
            Min=0

        if Hour>=24:
            Day+=1
            Hour=0
            
#----------------------Update meausrments------------------------

        SumArrTotEnergy+=np.multiply(ActArrTotPower,dt)
        SumArrGrdEnergy+=np.multiply(ActArrGrdPower,dt)
        
        hsb.updateBatteryValues(dt)

        for q in range (NumberOfCars):
            car[q].updateBatteryValues(dt)

#-------------------External time-------------------
        a=1
        while (StrTime+(NumSec*t)*nano>time.time_ns()):
            None

        
    else:

        Sec=0
        Min=0
        Hour=0
        Day=0

        AvgFlg=0
        StrFlg=False

        SystemRun=False
        SystemNeedEnergy=False

        EnergyMeter=0
        
        ReqArrPower=[0]*5
        SndArrPower=[0]*5
        GetArrPower=[0]*5
        ActArrPower=[0]*5

        ActArrGrdEnergy=[0]*5
        SumArrGrdPower=[0]*5
        ActArrTotEnergy=[0]*5
        SumArrTotPower=[0]*5

        if a==1:
            savingMeasurements.savingMeasurements(UserIndex,TimeOfTest,NumberOfCars)
            a=0

        print("System don't work")

        SystemRun=ethReg.getSystemRun()